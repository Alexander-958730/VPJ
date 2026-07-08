# apps/utils/export_utils.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
from io import BytesIO

FONT_NAME = 'Helvetica'


def export_complaints_to_excel(complaints, filename="denuncias.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Denuncias"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Titulo', 'Productor', 'Tipo', 'Descripcion', 'Estado', 'Fecha Creacion']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row, c in enumerate(complaints, 2):
        productor_nombre = f"{c.complained_against.first_name} {c.complained_against.last_name}".strip() or c.complained_against.username
        
        ws.cell(row=row, column=1, value=c.id)
        ws.cell(row=row, column=2, value=c.title or '')
        ws.cell(row=row, column=3, value=productor_nombre)
        ws.cell(row=row, column=4, value=c.get_complaint_type_display() if hasattr(c, 'get_complaint_type_display') else c.complaint_type)
        ws.cell(row=row, column=5, value=c.description[:200] if c.description else '')
        ws.cell(row=row, column=6, value=c.get_status_display() if hasattr(c, 'get_status_display') else c.status)
        ws.cell(row=row, column=7, value=c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '')
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_complaints_to_pdf(complaints, filename="denuncias.pdf"):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements.append(Paragraph("REPORTE DE DENUNCIAS", title_style))
    elements.append(Spacer(1, 20))
    
    data = [['ID', 'Titulo', 'Productor', 'Estado', 'Fecha', 'Descripcion']]
    
    for c in complaints[:100]:
        productor_nombre = f"{c.complained_against.first_name} {c.complained_against.last_name}".strip() or c.complained_against.username
        estado = c.get_status_display() if hasattr(c, 'get_status_display') else c.status
        
        data.append([
            str(c.id),
            c.title[:30] if c.title else '',
            productor_nombre[:25],
            estado,
            c.created_at.strftime('%d/%m/%Y') if c.created_at else '',
            (c.description[:80] + '...') if c.description and len(c.description) > 80 else (c.description or '')
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def export_productores_to_pdf(productores, filename="productores.pdf"):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    
    elements.append(Paragraph("LISTADO DE PRODUCTORES", title_style))
    elements.append(Paragraph(f"Total: {productores.count()} productores registrados", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    data = [['ID', 'Nombre', 'Email', 'Estado', 'Fecha Registro']]
    
    for p in productores:
        nombre_completo = f"{p.first_name} {p.last_name}".strip() or p.username
        estado = 'Aprobado' if p.is_approved else 'Pendiente'
        
        data.append([
            str(p.id),
            nombre_completo[:30],
            p.email[:25] if p.email else '',
            estado,
            p.date_joined.strftime('%d/%m/%Y') if p.date_joined else ''
        ])
    
    table = Table(data, repeatRows=1, colWidths=[40, 120, 100, 60, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response