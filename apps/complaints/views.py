"""
APPS.COMPLAINTS.VIEWS
=====================
Vistas para el sistema de denuncias, gestión de ONPECO, reportes y backups.

Este módulo contiene todas las vistas relacionadas con:
- Denuncias (creación, visualización, gestión)
- Portal ONPECO (dashboard, estadísticas)
- Backups y restauración del sistema
- Reportes de ventas y exportaciones
- Sistema de reputación de productores

Roles de usuario:
- consumidor: Puede crear denuncias, ver sus denuncias
- productor: Puede recibir denuncias, ver reportes de ventas
- regulador (ONPECO): Acceso completo a gestión de denuncias, portal, backups
- acopio (Centro de Acopio): Acceso limitado a supervisión
"""

import os
import datetime
import json
import zipfile
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.http import JsonResponse, Http404, HttpResponseForbidden, HttpResponse
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta
from functools import wraps
from .models import Complaint, ComplaintUpdate
from .services import EstadisticasONPECO
from core.utils import obtener_fechas


# =============================================================================
# DECORADOR: onpeco_required
# =============================================================================
# Propósito: Restringe el acceso a vistas SOLO para usuarios de ONPECO
# (staff o reguladores). Excluye al Centro de Acopio.
#
# Parámetros:
#   - view_func: La función de vista a decorar
#
# Retorna:
#   - wrapper: Función que verifica permisos antes de ejecutar la vista
#
# Usado por:
#   - Todas las vistas de administración de ONPECO
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#   - ✅ Staff (Django admin)
#   - ❌ Centro de Acopio
#   - ❌ Productor
#   - ❌ Consumidor
# =============================================================================
def onpeco_required(view_func):
    """Decorador que permite acceso solo a staff o reguladores ONPECO (excluye Centro de Acopio)"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            # Excluir al Centro de Acopio
            if request.user.username == 'centro_acopio' or getattr(request.user, 'role', '') == 'acopio':
                return HttpResponseForbidden("El Centro de Acopio no tiene acceso a esta sección.")
            if request.user.is_staff or getattr(request.user, 'role', '') == 'regulador':
                return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder")
    return wrapper


# =============================================================================
# FUNCIÓN: mis_denuncias
# =============================================================================
# Propósito: Muestra todas las denuncias creadas por el usuario actual
#
# Parámetros:
#   - request: Objeto HttpRequest con los datos de la sesión
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/mis_denuncias.html'
#     con la lista de denuncias del usuario
#
# URLs asociadas:
#   - /denuncias/mis-denuncias/
#
# Roles permitidos:
#   - ✅ Consumidor
#   - ✅ Productor (si tiene denuncias)
#   - ❌ Regulador (tiene su propia vista)
# =============================================================================
@login_required
def mis_denuncias(request):
    """Muestra las denuncias del usuario actual"""
    denuncias_usuario = Complaint.objects.filter(complainant=request.user).order_by('-created_at')
    return render(request, 'complaints/mis_denuncias.html', {'denuncias': denuncias_usuario})


# =============================================================================
# FUNCIÓN: detalle_denuncia
# =============================================================================
# Propósito: Muestra el detalle de una denuncia específica y permite a ONPECO
# actualizar su estado (pendiente, en investigación, resuelta, rechazada)
#
# Parámetros:
#   - request: Objeto HttpRequest
#   - pk: ID de la denuncia a mostrar
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/detalle_denuncia.html' con el detalle
#     y el historial de actualizaciones
#
# URLs asociadas:
#   - /denuncias/detalle/<pk>/
#
# Roles permitidos:
#   - ✅ Denunciante (creador de la denuncia)
#   - ✅ Denunciado (persona contra quien se denuncia)
#   - ✅ Regulador (ONPECO) - puede actualizar estado
#   - ❌ Otros usuarios
#
# Flujo de trabajo:
#   1. Verificar permisos del usuario
#   2. Si es POST y es ONPECO, actualizar estado
#   3. Crear registro de actualización (ComplaintUpdate)
#   4. Mostrar mensaje de éxito según el nuevo estado
# =============================================================================
@login_required
def detalle_denuncia(request, pk):
    """Muestra el detalle de una denuncia específica y permite actualizar estado"""
    denuncia_actual = get_object_or_404(Complaint, pk=pk)
    
    # Verificar permisos: solo el denunciante, el denunciado o ONPECO pueden ver
    if not (request.user == denuncia_actual.complainant or 
            request.user == denuncia_actual.complained_against or 
            request.user.is_staff or 
            getattr(request.user, 'role', '') == 'regulador'):
        return HttpResponseForbidden("No tienes permiso para ver esta denuncia")
    
    # Procesar actualización de estado (solo para ONPECO)
    if request.method == 'POST' and (request.user.is_staff or getattr(request.user, 'role', '') == 'regulador'):
        nuevo_estado = request.POST.get('status')
        comentario_actualizacion = request.POST.get('comentario', '')
        
        # Validar que el estado sea válido
        estados_validos = ['pending', 'investigating', 'resolved', 'rejected']
        
        if nuevo_estado in estados_validos:
            estado_anterior = denuncia_actual.status
            
            # Actualizar el estado
            denuncia_actual.status = nuevo_estado
            denuncia_actual.save()
            
            # Crear registro de actualización con los nombres correctos de los campos
            ComplaintUpdate.objects.create(
                complaint=denuncia_actual,
                old_status=estado_anterior,
                new_status=nuevo_estado,
                comment=comentario_actualizacion,
                created_by=request.user
            )
            
            # Mensaje de éxito según el nuevo estado
            if nuevo_estado == 'pending':
                messages.success(request, '✅ Denuncia marcada como Pendiente')
            elif nuevo_estado == 'investigating':
                messages.success(request, '✅ Denuncia en investigación - Se ha iniciado la investigación')
            elif nuevo_estado == 'resolved':
                messages.success(request, '✅ Denuncia Resuelta - Caso cerrado')
            elif nuevo_estado == 'rejected':
                messages.success(request, '✅ Denuncia Rechazada - No se encontraron evidencias suficientes')
            
            # Redirigir para evitar reenvío del formulario
            return redirect('complaints:detalle_denuncia', pk=denuncia_actual.id)
        else:
            messages.error(request, f'❌ Estado no válido: {nuevo_estado}')
    
    # Obtener historial de actualizaciones
    historial_updates = ComplaintUpdate.objects.filter(complaint=denuncia_actual).order_by('-created_at')
    
    context = {
        'denuncia': denuncia_actual,
        'updates': historial_updates,
    }
    return render(request, 'complaints/detalle_denuncia.html', context)


# =============================================================================
# FUNCIÓN: crear_denuncia
# =============================================================================
# Propósito: Permite a un usuario crear una nueva denuncia contra un producto
# o productor. Genera un ticket único de seguimiento.
#
# Parámetros:
#   - request: Objeto HttpRequest
#   - producto_id: (Opcional) ID del producto denunciado
#
# Retorna:
#   - GET: Renderiza 'complaints/crear_denuncia.html' con el formulario
#   - POST: Crea la denuncia y redirige a 'mis_denuncias'
#
# URLs asociadas:
#   - /denuncias/crear/
#   - /denuncias/crear/<producto_id>/
#
# Roles permitidos:
#   - ✅ Consumidor (principalmente)
#   - ✅ Productor (si tiene una queja)
#   - ❌ Regulador (tiene otras herramientas)
#
# Flujo de trabajo:
#   1. Si se recibe producto_id, obtener el producto y su vendedor
#   2. Procesar formulario POST con los datos de la denuncia
#   3. Generar ticket único (CD-XXXXXX)
#   4. Guardar denuncia en estado 'pending'
#   5. Mostrar mensaje con el número de ticket
# =============================================================================
@login_required
def crear_denuncia(request, producto_id=None):
    """
    Crear una denuncia contra un producto o productor
    """
    from apps.marketplace.models import Product
    from apps.users.models import User
    
    producto_denunciado = None
    productor_denunciado = None
    
    # Si se pasó un producto_id, obtener el producto y su productor
    if producto_id:
        producto_denunciado = get_object_or_404(Product, id=producto_id)
        productor_denunciado = getattr(producto_denunciado, 'vendedor', None) or getattr(producto_denunciado, 'productor', None)
    
    if request.method == 'POST':
        titulo_denuncia = request.POST.get('titulo')
        tipo_denuncia = request.POST.get('tipo')
        prioridad_denuncia = request.POST.get('prioridad')
        descripcion_denuncia = request.POST.get('descripcion')
        
        # Generar ticket único
        ultima_denuncia = Complaint.objects.order_by('-id').first()
        if ultima_denuncia and ultima_denuncia.ticket_number:
            try:
                numero_ticket = int(ultima_denuncia.ticket_number.split('-')[1]) + 1
            except:
                numero_ticket = 1
        else:
            numero_ticket = 1
        
        ticket_generado = f"CD-{numero_ticket:06d}"
        
        denuncia_creada = Complaint.objects.create(
            ticket_number=ticket_generado,
            title=titulo_denuncia,
            complaint_type=tipo_denuncia,
            priority=prioridad_denuncia,
            description=descripcion_denuncia,
            complainant=request.user,
            complained_against=productor_denunciado if productor_denunciado else None,
            producto=producto_denunciado,
            status='pending'
        )
        
        messages.success(request, f'✅ Denuncia creada. Tu número de seguimiento es: {ticket_generado}')
        return redirect('complaints:mis_denuncias')
    
    context = {
        'producto': producto_denunciado,
        'productor': productor_denunciado,
    }
    return render(request, 'complaints/crear_denuncia.html', context)


# =============================================================================
# FUNCIÓN: lista_denuncias
# =============================================================================
# Propósito: Panel de gestión de denuncias para ONPECO. Permite buscar
# denuncias por los últimos dígitos del ticket.
#
# Parámetros:
#   - request: Objeto HttpRequest con parámetro GET 'busqueda'
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/lista_denuncias.html' con las
#     denuncias clasificadas por estado
#
# URLs asociadas:
#   - /denuncias/lista/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#   - ❌ Centro de Acopio
#   - ❌ Consumidor
#   - ❌ Productor
#
# Flujo de trabajo:
#   1. Obtener término de búsqueda (últimos 1-4 dígitos del ticket)
#   2. Filtrar denuncias que terminen con esos dígitos
#   3. Clasificar por estado: pendientes, investigando, resueltas, rechazadas
#   4. Mostrar estadísticas en consola para debugging
# =============================================================================
@onpeco_required
def lista_denuncias(request):
    """
    Vista para que ONPECO gestione todas las denuncias del sistema.
    Permite búsqueda por últimos 1 a 4 dígitos del ticket.
    """
    # Obtener el parámetro de búsqueda
    busqueda_ticket = request.GET.get('busqueda', '').strip()
    
    # Obtener TODAS las denuncias ordenadas por fecha descendente
    todas_las_denuncias = Complaint.objects.all().order_by('-created_at')
    
    # Filtrar por últimos dígitos del ticket
    if busqueda_ticket and len(busqueda_ticket) >= 1:
        todas_las_denuncias = todas_las_denuncias.filter(
            ticket_number__endswith=busqueda_ticket
        )
        
        if not todas_las_denuncias.exists():
            messages.info(request, f'🔍 No se encontraron denuncias que terminen en "{busqueda_ticket}"')
    
    # Clasificar por estado
    denuncias_pendientes = todas_las_denuncias.filter(status='pending')
    denuncias_investigando = todas_las_denuncias.filter(status='investigating')
    denuncias_resueltas = todas_las_denuncias.filter(status='resolved')
    denuncias_rechazadas = todas_las_denuncias.filter(status='rejected')
    
    # Debug en consola
    print(f"📊 Estadísticas de denuncias - ONPECO:")
    print(f"   Pendientes: {denuncias_pendientes.count()}")
    print(f"   En investigación: {denuncias_investigando.count()}")
    print(f"   Resueltas: {denuncias_resueltas.count()}")
    print(f"   Rechazadas: {denuncias_rechazadas.count()}")
    print(f"   Total: {todas_las_denuncias.count()}")
    if busqueda_ticket:
        print(f"   🔍 Búsqueda: '{busqueda_ticket}'")
    
    context = {
        'pendientes': denuncias_pendientes,
        'investigando': denuncias_investigando,
        'resueltas': denuncias_resueltas,
        'rechazadas': denuncias_rechazadas,
        'busqueda': busqueda_ticket,
    }
    return render(request, 'complaints/lista_denuncias.html', context)


# =============================================================================
# FUNCIÓN: actualizar_denuncia
# =============================================================================
# Propósito: Redirige al detalle de una denuncia para actualizarla.
# Es un atajo para facilitar la navegación.
#
# Parámetros:
#   - request: Objeto HttpRequest
#   - pk: ID de la denuncia
#
# Retorna:
#   - HttpResponse: Redirige a 'detalle_denuncia'
#
# URLs asociadas:
#   - /denuncias/actualizar/<pk>/
# =============================================================================
@onpeco_required
def actualizar_denuncia(request, pk):
    """Redirige al detalle de la denuncia para actualizarla"""
    return redirect('complaints:detalle_denuncia', pk=pk)


# =============================================================================
# FUNCIÓN: gestion_backups
# =============================================================================
# Propósito: Muestra la interfaz de gestión de backups con la lista de
# archivos de respaldo disponibles y estadísticas.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/gestion_backups.html' con la lista
#     de backups y estadísticas
#
# URLs asociadas:
#   - /denuncias/gestion-backups/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#   - ❌ Centro de Acopio
#   - ❌ Consumidor
#   - ❌ Productor
#
# Estructura de archivos:
#   - backup_YYYY-MM-DD-HHMMSS.json: Backup de base de datos
#   - backup_media_YYYY-MM-DD-HHMMSS.zip: Backup de archivos multimedia
# =============================================================================
@onpeco_required
def gestion_backups(request):
    backups_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backups_dir):
        os.makedirs(backups_dir)
    
    # Listar backups
    lista_backups = []
    if os.path.exists(backups_dir):
        for archivo in os.listdir(backups_dir):
            ruta_archivo = os.path.join(backups_dir, archivo)
            if os.path.isfile(ruta_archivo):
                info_archivo = os.stat(ruta_archivo)
                lista_backups.append({
                    'nombre': archivo,
                    'fecha': datetime.datetime.fromtimestamp(info_archivo.st_mtime),
                    'tamano': round(info_archivo.st_size / 1024 / 1024, 2),
                })
        lista_backups.sort(key=lambda x: x['fecha'], reverse=True)
    
    # Calcular estadísticas de backups
    backups_manuales = 0
    backups_automaticos = 0
    backups_media = 0
    backups_dump = 0
    backups_otros = 0
    
    for backup_actual in lista_backups:
        nombre_backup = backup_actual['nombre']
        if 'punto_restauracion' in nombre_backup:
            backups_manuales += 1
        elif 'backup_auto' in nombre_backup:
            backups_automaticos += 1
        elif 'backup_media' in nombre_backup:
            backups_media += 1
        elif nombre_backup.endswith('.zip'):
            backups_media += 1
        elif nombre_backup.endswith('.dump'):
            backups_dump += 1
        elif 'backup_' in nombre_backup and '.json' in nombre_backup:
            backups_manuales += 1
        else:
            backups_otros += 1
    
    estadisticas_backups = {
        'total': len(lista_backups),
        'manual': backups_manuales,
        'auto': backups_automaticos,
        'media': backups_media,
        'dump': backups_dump,
        'otro': backups_otros,
    }
    
    return render(request, 'complaints/gestion_backups.html', {
        'backups': lista_backups,
        'stats': estadisticas_backups,
    })


# =============================================================================
# FUNCIÓN: crear_punto_restauracion
# =============================================================================
# Propósito: Crea un backup completo del sistema en dos archivos:
#   1. JSON: Exporta toda la base de datos (productos, usuarios, denuncias)
#   2. ZIP: Comprime todos los archivos de la carpeta media/
#
# Parámetros:
#   - request: Objeto HttpRequest con POST (descripción del backup)
#
# Retorna:
#   - HttpResponse: Redirige a 'gestion_backups' con mensaje de éxito/error
#
# URLs asociadas:
#   - /denuncias/gestion-backups/crear-punto/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#
# Estructura del backup JSON:
#   {
#       "version": "2.0",
#       "fecha": "2026-07-06-184530",
#       "tipo": "backup_completo",
#       "creado_por": "onpeco_regulador",
#       "estadisticas": { "total_productos": 45, "total_usuarios": 30 },
#       "datos": { "productos": [...], "usuarios": [...], "denuncias": [...] }
#   }
# =============================================================================
@onpeco_required
def crear_punto_restauracion(request):
    """Crea un backup completo del sistema: JSON (DB) + ZIP (Media)"""
    if request.method == 'POST':
        try:
            from apps.marketplace.models import Product
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            backups_dir = os.path.join(settings.BASE_DIR, 'backups')
            if not os.path.exists(backups_dir):
                os.makedirs(backups_dir)
            
            timestamp_backup = datetime.datetime.now().strftime('%Y-%m-%d-%H%M%S')
            descripcion_backup = request.POST.get('descripcion', 'Sin descripción')
            
            # 1. CREAR BACKUP JSON (BASE DE DATOS)
            archivo_json = os.path.join(backups_dir, f'backup_{timestamp_backup}.json')
            
            lista_productos = list(Product.objects.all().values())
            lista_usuarios = list(User.objects.all().values('id', 'username', 'email', 'first_name', 'last_name', 'is_staff', 'is_active', 'date_joined', 'last_login', 'role', 'is_approved', 'phone', 'business_name', 'address'))
            lista_denuncias = list(Complaint.objects.all().values())
            
            lista_updates_denuncias = []
            try:
                lista_updates_denuncias = list(ComplaintUpdate.objects.all().values())
            except:
                pass
            
            datos_backup = {
                'version': '2.0',
                'fecha': timestamp_backup,
                'fecha_legible': datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'tipo': 'backup_completo',
                'creado_por': request.user.username,
                'creado_por_id': request.user.id,
                'descripcion': descripcion_backup,
                'estadisticas': {
                    'total_productos': len(lista_productos),
                    'total_usuarios': len(lista_usuarios),
                    'total_denuncias': len(lista_denuncias),
                    'total_actualizaciones': len(lista_updates_denuncias),
                },
                'datos': {
                    'productos': lista_productos,
                    'usuarios': lista_usuarios,
                    'denuncias': lista_denuncias,
                    'denuncias_updates': lista_updates_denuncias,
                }
            }
            
            with open(archivo_json, 'w', encoding='utf-8') as f:
                json.dump(datos_backup, f, indent=2, ensure_ascii=False, default=str)
            
            tamanio_json = os.path.getsize(archivo_json) / 1024
            
            # 2. CREAR BACKUP ZIP (ARCHIVOS MEDIA)
            archivo_zip = os.path.join(backups_dir, f'backup_media_{timestamp_backup}.zip')
            directorio_media = os.path.join(settings.BASE_DIR, 'media')
            
            with zipfile.ZipFile(archivo_zip, 'w', zipfile.ZIP_DEFLATED) as zip_backup:
                if os.path.exists(directorio_media):
                    for root, dirs, files in os.walk(directorio_media):
                        for file in files:
                            ruta_archivo_media = os.path.join(root, file)
                            nombre_relativo = os.path.relpath(ruta_archivo_media, settings.BASE_DIR)
                            zip_backup.write(ruta_archivo_media, nombre_relativo)
            
            tamanio_zip = os.path.getsize(archivo_zip) / 1024
            
            # 3. MENSAJE DE ÉXITO CON AMBOS ARCHIVOS
            stats_backup = datos_backup['estadisticas']
            messages.success(
                request, 
                f'✅ Backup completo creado:\n'
                f'   📄 JSON: backup_{timestamp_backup}.json ({tamanio_json:.1f} KB)\n'
                f'   📦 ZIP: backup_media_{timestamp_backup}.zip ({tamanio_zip:.1f} KB)\n'
                f'   📝 Descripción: {descripcion_backup}\n'
                f'   📊 {stats_backup["total_productos"]} productos, {stats_backup["total_usuarios"]} usuarios, {stats_backup["total_denuncias"]} denuncias'
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear backup completo: {str(e)}')
        
        return redirect('complaints:gestion_backups')
    
    return redirect('complaints:gestion_backups')


# =============================================================================
# FUNCIÓN: restaurar_backup
# =============================================================================
# Propósito: Restaura completamente el sistema desde un backup:
#   1. Base de datos desde archivo JSON
#   2. Archivos multimedia desde archivo ZIP (si existe)
#
# Parámetros:
#   - request: Objeto HttpRequest
#   - backup_filename: Nombre del archivo JSON a restaurar
#
# Retorna:
#   - HttpResponse: Redirige a 'gestion_backups' con mensaje de éxito/error
#
# URLs asociadas:
#   - /denuncias/backups/restaurar/<backup_filename>/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#
# ADVERTENCIA:
#   - Esta operación SOBREESCRIBE la base de datos actual
#   - Se recomienda hacer un backup antes de restaurar
# =============================================================================
@onpeco_required
def restaurar_backup(request, backup_filename):
    """
    Restaura COMPLETAMENTE el sistema:
    1. Base de datos desde JSON
    2. Archivos multimedia desde ZIP (si existe)
    """
    import traceback
    
    try:
        backups_dir = os.path.join(settings.BASE_DIR, 'backups')
        ruta_backup = os.path.join(backups_dir, backup_filename)
        
        if not os.path.exists(ruta_backup):
            messages.error(request, f'❌ Archivo no encontrado: {backup_filename}')
            return redirect('complaints:gestion_backups')
        
        # Leer el archivo de respaldo
        with open(ruta_backup, 'r', encoding='utf-8', errors='ignore') as f:
            datos_backup = json.load(f)
        
        # Importar modelos
        from apps.marketplace.models import Product
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # 1. RESTAURAR BASE DE DATOS (JSON)
        
        # Caso 1: Formato antiguo (lista de productos)
        if isinstance(datos_backup, list):
            productos_restaurados = 0
            productos_omitidos = 0
            
            for producto_data in datos_backup:
                if isinstance(producto_data, dict) and 'name' in producto_data:
                    existe = Product.objects.filter(name=producto_data['name']).exists()
                    if not existe:
                        Product.objects.create(
                            name=producto_data['name'],
                            productor_id=producto_data.get('productor_id', 1),
                            description=producto_data.get('description', ''),
                            category=producto_data.get('category', 'otros'),
                            price=producto_data.get('price', 0),
                            unit=producto_data.get('unit', 'unidad'),
                            stock=producto_data.get('stock', 0),
                            stock_minimo=producto_data.get('stock_minimo', 0),
                            image=producto_data.get('image', ''),
                            available=producto_data.get('available', True),
                        )
                        productos_restaurados += 1
                    else:
                        productos_omitidos += 1
            
            messages.success(request, f'✅ Restauración completada: {productos_restaurados} productos restaurados, {productos_omitidos} omitidos')
            
            # Restaurar archivos media desde ZIP
            nombre_base = backup_filename.replace('.json', '').replace('backup_', '')
            nombre_zip = f'backup_media_{nombre_base}.zip'
            ruta_zip = os.path.join(backups_dir, nombre_zip)
            
            if os.path.exists(ruta_zip):
                with zipfile.ZipFile(ruta_zip, 'r') as zip_restaurar:
                    zip_restaurar.extractall(settings.BASE_DIR)
                messages.success(request, f'✅ Archivos multimedia restaurados desde {nombre_zip}')
            else:
                messages.warning(request, f'⚠️ No se encontró ZIP para {backup_filename}. Solo se restauró la base de datos.')
            
            return redirect('complaints:gestion_backups')
        
        # Caso 2: Formato nuevo con estructura 'datos'
        if isinstance(datos_backup, dict) and 'datos' in datos_backup:
            usuarios_restaurados = 0
            usuarios_omitidos = 0
            productos_restaurados = 0
            productos_omitidos = 0
            denuncias_restauradas = 0
            denuncias_omitidas = 0
            
            # Restaurar usuarios
            if 'usuarios' in datos_backup['datos']:
                for usuario_data in datos_backup['datos']['usuarios']:
                    if isinstance(usuario_data, dict) and 'username' in usuario_data:
                        existe = User.objects.filter(username=usuario_data['username']).exists()
                        if not existe:
                            usuario_nuevo = User(
                                username=usuario_data['username'],
                                email=usuario_data.get('email', ''),
                                first_name=usuario_data.get('first_name', ''),
                                last_name=usuario_data.get('last_name', ''),
                                is_staff=usuario_data.get('is_staff', False),
                                is_active=usuario_data.get('is_active', True),
                            )
                            usuario_nuevo.set_unusable_password()
                            usuario_nuevo.save()
                            usuarios_restaurados += 1
                        else:
                            usuarios_omitidos += 1
            
            # Restaurar productos
            if 'productos' in datos_backup['datos']:
                for producto_data in datos_backup['datos']['productos']:
                    if isinstance(producto_data, dict) and 'name' in producto_data:
                        existe = Product.objects.filter(name=producto_data['name']).exists()
                        if not existe:
                            Product.objects.create(
                                name=producto_data['name'],
                                productor_id=producto_data.get('productor_id', 1),
                                description=producto_data.get('description', ''),
                                category=producto_data.get('category', 'otros'),
                                price=producto_data.get('price', 0),
                                unit=producto_data.get('unit', 'unidad'),
                                stock=producto_data.get('stock', 0),
                                stock_minimo=producto_data.get('stock_minimo', 0),
                                image=producto_data.get('image', ''),
                                available=producto_data.get('available', True),
                            )
                            productos_restaurados += 1
                        else:
                            productos_omitidos += 1
            
            # Restaurar denuncias
            if 'denuncias' in datos_backup['datos']:
                for denuncia_data in datos_backup['datos']['denuncias']:
                    if isinstance(denuncia_data, dict):
                        ticket = denuncia_data.get('ticket_number', '')
                        existe = Complaint.objects.filter(ticket_number=ticket).exists() if ticket else False
                        if not existe:
                            Complaint.objects.create(
                                complainant_id=denuncia_data.get('complainant_id', 1),
                                complained_against_id=denuncia_data.get('complained_against_id', 1),
                                ticket_number=ticket,
                                title=denuncia_data.get('title', ''),
                                complaint_type=denuncia_data.get('complaint_type', ''),
                                description=denuncia_data.get('description', ''),
                                status=denuncia_data.get('status', 'pending'),
                                priority=denuncia_data.get('priority', 'medium'),
                                created_at=denuncia_data.get('created_at', timezone.now()),
                                updated_at=denuncia_data.get('updated_at', timezone.now()),
                            )
                            denuncias_restauradas += 1
                        else:
                            denuncias_omitidas += 1
            
            mensaje_restauracion = (
                f'✅ Base de datos restaurada:\n'
                f'   👤 Usuarios: {usuarios_restaurados} restaurados, {usuarios_omitidos} omitidos\n'
                f'   📦 Productos: {productos_restaurados} restaurados, {productos_omitidos} omitidos\n'
                f'   📋 Denuncias: {denuncias_restauradas} restauradas, {denuncias_omitidas} omitidas'
            )
            messages.success(request, mensaje_restauracion)
            
            # Restaurar archivos media desde ZIP
            nombre_base = backup_filename.replace('.json', '').replace('backup_', '')
            nombre_zip = f'backup_media_{nombre_base}.zip'
            ruta_zip = os.path.join(backups_dir, nombre_zip)
            
            if os.path.exists(ruta_zip):
                with zipfile.ZipFile(ruta_zip, 'r') as zip_restaurar:
                    zip_restaurar.extractall(settings.BASE_DIR)
                messages.success(request, f'✅ Archivos multimedia restaurados desde {nombre_zip}')
            else:
                messages.warning(request, f'⚠️ No se encontró ZIP para {backup_filename}. Solo se restauró la base de datos.')
            
            return redirect('complaints:gestion_backups')
        
        messages.error(request, f'❌ Formato de backup no reconocido')
        
    except Exception as e:
        messages.error(request, f'❌ Error al restaurar: {str(e)}')
    
    return redirect('complaints:gestion_backups')


# =============================================================================
# FUNCIÓN: denuncias_por_mes_api
# =============================================================================
# Propósito: API para el gráfico de denuncias por mes. Devuelve datos en
# formato JSON para ser consumidos por Chart.js.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - JsonResponse: Datos con labels (meses) y datos (cantidad de denuncias)
#
# URLs asociadas:
#   - /denuncias/api/denuncias-por-mes/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#
# Formato de respuesta:
#   {
#       "labels": ["Ene 2026", "Feb 2026", ...],
#       "datos": [5, 8, 12, ...],
#       "titulo": "Denuncias por mes (último año)"
#   }
# =============================================================================
@onpeco_required
def denuncias_por_mes_api(request):
    """API para gráfico de denuncias por mes (solo ONPECO regulador)"""
    fecha_limite = timezone.now() - timedelta(days=365)
    
    denuncias_por_mes = (
        Complaint.objects.filter(created_at__gte=fecha_limite)
        .annotate(mes=TruncMonth('created_at'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )
    
    labels_meses = []
    datos_meses = []
    
    for item in denuncias_por_mes:
        if item['mes']:
            labels_meses.append(item['mes'].strftime('%b %Y'))
            datos_meses.append(item['total'])
    
    return JsonResponse({
        'labels': labels_meses,
        'datos': datos_meses,
        'titulo': 'Denuncias por mes (último año)'
    })


# =============================================================================
# FUNCIÓN: reporte_denuncias
# =============================================================================
# Propósito: Vista principal del reporte de denuncias con gráfico.
# Renderiza la página que contiene el gráfico de Chart.js.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/reporte_denuncias.html'
#
# URLs asociadas:
#   - /denuncias/reporte-denuncias/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def reporte_denuncias(request):
    """Vista principal del reporte de denuncias con gráfico"""
    return render(request, 'complaints/reporte_denuncias.html')


# =============================================================================
# FUNCIÓN: portal_onpeco
# =============================================================================
# Propósito: Dashboard principal de ONPECO. Muestra estadísticas clave del
# sistema en tarjetas visuales: total de denuncias, productos, usuarios, etc.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Renderiza 'onpeco/portal.html' con todas las estadísticas
#
# URLs asociadas:
#   - /denuncias/portal/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#   - ❌ Centro de Acopio (bloqueado)
#
# Métricas mostradas:
#   - Denuncias: total, pendientes, investigando, aprobadas, rechazadas
#   - Productos: total en el marketplace
#   - Usuarios: total, productores, consumidores
# =============================================================================
from apps.marketplace.models import Product
from django.contrib.auth import get_user_model
User = get_user_model()

@login_required
def portal_onpeco(request):
    """Portal exclusivo para ONPECO con todos los gráficos y gestión"""
    
    # Bloquear al Centro de Acopio
    if request.user.username == 'centro_acopio' or getattr(request.user, 'role', '') == 'acopio':
        messages.error(request, '❌ El Centro de Acopio no tiene acceso al portal de ONPECO.')
        return redirect('inicio')
    
    if not request.user.is_staff and getattr(request.user, 'role', '') != 'regulador':
        return HttpResponseForbidden("No tienes permiso para acceder al portal de ONPECO.")
    
    # Obtener estadísticas desde el servicio
    estadisticas_resumen = EstadisticasONPECO.get_resumen_general()
    
    context = {
        'total_denuncias': estadisticas_resumen['total_denuncias'],
        'denuncias_pendientes': estadisticas_resumen['denuncias_pendientes'],
        'denuncias_investigando': estadisticas_resumen['denuncias_investigando'],
        'denuncias_aprobadas': estadisticas_resumen['denuncias_aprobadas'],
        'denuncias_rechazadas': estadisticas_resumen['denuncias_rechazadas'],
        'total_productos': estadisticas_resumen['total_productos'],
        'total_usuarios': estadisticas_resumen['total_usuarios'],
        'total_productores': estadisticas_resumen['total_productores'],
        'total_consumidores': estadisticas_resumen['total_consumidores'],
        'user': request.user,
    }
    
    return render(request, 'onpeco/portal.html', context)


# =============================================================================
# FUNCIÓN: api_denuncias_recientes
# =============================================================================
# Propósito: API para obtener las últimas 10 denuncias registradas.
# Utilizada para el dashboard de ONPECO y para actualizaciones en tiempo real.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - JsonResponse: Lista de las últimas denuncias
#
# URLs asociadas:
#   - /denuncias/api/recientes/
#
# Roles permitidos:
#   - ✅ Usuarios autenticados
# =============================================================================
@login_required
def api_denuncias_recientes(request):
    """API para obtener las últimas denuncias"""
    denuncias_recientes = Complaint.objects.order_by('-created_at')[:10]
    data_json = []
    for denuncia_actual in denuncias_recientes:
        data_json.append({
            'ticket_number': denuncia_actual.ticket_number,
            'title': denuncia_actual.title,
            'status': denuncia_actual.status,
            'created_at': denuncia_actual.created_at.strftime('%d/%m/%Y') if denuncia_actual.created_at else '',
        })
    return JsonResponse(data_json, safe=False)


# =============================================================================
# FUNCIÓN: productores_mas_denunciados
# =============================================================================
# Propósito: Muestra un ranking de los productores con mayor cantidad de
# denuncias registradas en su contra.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/productores_mas_denunciados.html'
#
# URLs asociadas:
#   - /denuncias/productores-mas-denunciados/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def productores_mas_denunciados(request):
    """
    Vista para el portal ONPECO que muestra los productores 
    con mayor cantidad de denuncias registradas.
    """
    # Usar el servicio de estadísticas
    productores_top = EstadisticasONPECO.get_top_productores_denunciados(limite=20)
    total_denuncias_top = sum([productor_actual.total_denuncias for productor_actual in productores_top])
    
    context = {
        'productores_denunciados': productores_top,
        'total_denuncias': total_denuncias_top,
    }
    return render(request, 'complaints/productores_mas_denunciados.html', context)


# =============================================================================
# FUNCIÓN: gestion_reputacion
# =============================================================================
# Propósito: Muestra la interfaz de gestión de reputación de productores
# y suplidores. Permite a ONPECO ver y asignar niveles de reputación.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/gestion_reputacion.html'
#
# URLs asociadas:
#   - /denuncias/gestion-reputacion/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def gestion_reputacion(request):
    """
    Vista para que ONPECO gestione la reputación de productores y suplidores
    """
    from apps.reviews.models import Review
    from django.db.models import Avg
    
    usuarios_reputacion = User.objects.filter(
        Q(role='productor') | Q(role='suplidor'),
        is_approved=True
    ).order_by('business_name')
    
    for usuario_actual in usuarios_reputacion:
        usuario_actual.total_reseñas = Review.objects.filter(productor=usuario_actual).count()
        promedio_calificacion = Review.objects.filter(productor=usuario_actual).aggregate(avg=Avg('rating'))['avg']
        usuario_actual.promedio_calificacion = round(promedio_calificacion, 1) if promedio_calificacion else 0
        usuario_actual.total_denuncias = Complaint.objects.filter(complained_against=usuario_actual).count()
        
        try:
            from .models import Reputacion
            ultima_reputacion = Reputacion.objects.filter(productor=usuario_actual).first()
            usuario_actual.ultima_reputacion = ultima_reputacion
        except:
            usuario_actual.ultima_reputacion = None
    
    context = {
        'usuarios': usuarios_reputacion,
    }
    return render(request, 'complaints/gestion_reputacion.html', context)


# =============================================================================
# FUNCIÓN: asignar_reputacion
# =============================================================================
# Propósito: Permite a ONPECO asignar un nivel de reputación a un productor
# o suplidor específico, con un comentario justificativo.
#
# Parámetros:
#   - request: Objeto HttpRequest (GET, POST)
#   - user_id: ID del usuario a calificar
#
# Retorna:
#   - GET: Renderiza 'complaints/asignar_reputacion.html' con el formulario
#   - POST: Asigna la reputación y redirige a 'gestion_reputacion'
#
# URLs asociadas:
#   - /denuncias/asignar-reputacion/<user_id>/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def asignar_reputacion(request, user_id):
    """
    Vista para asignar reputación a un productor/suplidor
    """
    usuario_reputacion = get_object_or_404(User, id=user_id, role__in=['productor', 'suplidor'])
    
    if request.method == 'POST':
        nivel_reputacion = request.POST.get('nivel')
        comentario_reputacion = request.POST.get('comentario')
        
        if not nivel_reputacion or not comentario_reputacion:
            messages.error(request, '❌ Debes seleccionar un nivel y escribir un comentario.')
            return redirect('complaints:gestion_reputacion')
        
        from .models import Reputacion
        reputacion_creada = Reputacion.objects.create(
            productor=usuario_reputacion,
            nivel=nivel_reputacion,
            comentario=comentario_reputacion,
            creado_por=request.user
        )
        
        usuario_reputacion.reputacion_actual = nivel_reputacion
        usuario_reputacion.save()
        
        messages.success(request, f'✅ Reputación asignada a {usuario_reputacion.business_name}: {reputacion_creada.get_nivel_display()}')
        return redirect('complaints:gestion_reputacion')
    
    from .models import Reputacion
    context = {
        'usuario': usuario_reputacion,
        'niveles': Reputacion.NIVELES_REPUTACION,
    }
    return render(request, 'complaints/asignar_reputacion.html', context)


# =============================================================================
# FUNCIÓN: historial_reputacion
# =============================================================================
# Propósito: Muestra el historial completo de cambios de reputación de un
# productor o suplidor específico.
#
# Parámetros:
#   - request: Objeto HttpRequest
#   - user_id: ID del usuario a consultar
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/historial_reputacion.html'
#
# URLs asociadas:
#   - /denuncias/historial-reputacion/<user_id>/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def historial_reputacion(request, user_id):
    """
    Vista para ver el historial completo de reputación de un usuario
    """
    usuario_reputacion = get_object_or_404(User, id=user_id, role__in=['productor', 'suplidor'])
    from .models import Reputacion
    historial_reputacion = Reputacion.objects.filter(productor=usuario_reputacion).order_by('-created_at')
    
    context = {
        'usuario': usuario_reputacion,
        'historial': historial_reputacion,
    }
    return render(request, 'complaints/historial_reputacion.html', context)


# =============================================================================
# REPORTES DE VENTAS PARA ONPECO
# =============================================================================

from apps.marketplace.models import Product as Producto
from apps.cart.models import Order as Pedido, OrderItem as DetallePedido
from apps.users.models import User as Usuario
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# =============================================================================
# FUNCIÓN: reporte_ventas_general
# =============================================================================
# Propósito: Genera un reporte general de ventas por período. Muestra un
# resumen de ventas por vendedor (productor o suplidor).
#
# Parámetros:
#   - request: Objeto HttpRequest con GET (fecha_inicio, fecha_fin)
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/reporte_ventas_general.html'
#
# URLs asociadas:
#   - /denuncias/reportes/ventas/general/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
#
# Nota: Cada pedido cuenta como 1 venta (transacción)
# =============================================================================
@onpeco_required
def reporte_ventas_general(request):
    """Vista para reporte general de ventas por período (incluye productores y suplidores)"""
    
    # Obtener fechas del filtro
    fecha_inicio, fecha_fin = obtener_fechas(request)
    
    # Obtener pedidos entregados en el período
    pedidos_del_periodo = Pedido.objects.filter(
        created_at__date__gte=fecha_inicio,
        created_at__date__lte=fecha_fin,
        status='delivered'
    )
    
    # Ventas por vendedor
    pedidos_contados_por_vendedor = {}
    ventas_por_vendedor = {}
    
    for pedido_actual in pedidos_del_periodo:
        vendedores_en_pedido = set()
        for detalle_pedido in pedido_actual.items.all():
            vendedor_del_producto = detalle_pedido.product.vendedor
            vendedor_id = vendedor_del_producto.id
            vendedores_en_pedido.add(vendedor_id)
            
            if vendedor_id not in ventas_por_vendedor:
                ventas_por_vendedor[vendedor_id] = {
                    'nombre_vendedor': vendedor_del_producto.business_name or vendedor_del_producto.username,
                    'rol_vendedor': vendedor_del_producto.role,
                    'cantidad_transacciones': 0,
                    'total_ingresos': 0,
                    'productos_vendidos': []
                }
            
            ventas_por_vendedor[vendedor_id]['total_ingresos'] += float(detalle_pedido.price * detalle_pedido.quantity)
            ventas_por_vendedor[vendedor_id]['productos_vendidos'].append(detalle_pedido.product.name)
        
        # Contar 1 venta por vendedor en este pedido
        for vendedor_id in vendedores_en_pedido:
            clave_pedido_vendedor = f"{pedido_actual.id}_{vendedor_id}"
            if clave_pedido_vendedor not in pedidos_contados_por_vendedor:
                pedidos_contados_por_vendedor[clave_pedido_vendedor] = True
                ventas_por_vendedor[vendedor_id]['cantidad_transacciones'] += 1
    
    # Calcular totales
    total_transacciones_sistema = 0
    for datos_vendedor in ventas_por_vendedor.values():
        total_transacciones_sistema += datos_vendedor['cantidad_transacciones']
    
    total_ingresos_sistema = 0
    for pedido_actual in pedidos_del_periodo:
        for detalle_pedido in pedido_actual.items.all():
            total_ingresos_sistema += float(detalle_pedido.price * detalle_pedido.quantity)
    
    ventas_ordenadas = sorted(
        ventas_por_vendedor.items(),
        key=lambda x: x[1]['total_ingresos'],
        reverse=True
    )
    
    total_productores_activos = sum(1 for datos_vendedor in ventas_por_vendedor.values() if datos_vendedor['rol_vendedor'] == 'productor')
    total_suplidores_activos = sum(1 for datos_vendedor in ventas_por_vendedor.values() if datos_vendedor['rol_vendedor'] == 'suplidor')
    
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_ventas': total_transacciones_sistema,
        'total_ingresos': total_ingresos_sistema,
        'ventas_ordenadas': ventas_ordenadas,
        'total_productores': total_productores_activos,
        'total_suplidores': total_suplidores_activos,
        'pedidos': pedidos_del_periodo,
    }
    return render(request, 'complaints/reporte_ventas_general.html', context)


# =============================================================================
# FUNCIÓN: reporte_ventas_productor
# =============================================================================
# Propósito: Genera un reporte detallado de ventas de un productor específico.
# Agrupa los productos por pedido.
#
# Parámetros:
#   - request: Objeto HttpRequest con GET (fecha_inicio, fecha_fin)
#   - productor_id: ID del productor a consultar
#
# Retorna:
#   - HttpResponse: Renderiza 'complaints/reporte_ventas_productor.html'
#
# URLs asociadas:
#   - /denuncias/reportes/ventas/productor/<productor_id>/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def reporte_ventas_productor(request, productor_id):
    """Vista para reporte de ventas de un productor específico con detalle agrupado por pedido"""
    
    productor_seleccionado = get_object_or_404(Usuario, id=productor_id, role='productor')
    
    # Obtener fechas del filtro
    fecha_inicio, fecha_fin = obtener_fechas(request)
    
    # Obtener pedidos del productor en el período
    pedidos_del_productor = Pedido.objects.filter(
        items__product__vendedor=productor_seleccionado,
        created_at__date__gte=fecha_inicio,
        created_at__date__lte=fecha_fin,
        status='delivered'
    ).distinct().order_by('-created_at')
    
    # Estructura: Ventas agrupadas por pedido
    ventas_agrupadas = []
    total_ingresos_productor = 0
    total_ventas_productor = pedidos_del_productor.count()
    
    for pedido_actual in pedidos_del_productor:
        items_del_productor_en_pedido = []
        subtotal_pedido_actual = 0
        
        for detalle_pedido in pedido_actual.items.all():
            if detalle_pedido.product.vendedor == productor_seleccionado:
                subtotal_detalle = detalle_pedido.price * detalle_pedido.quantity
                items_del_productor_en_pedido.append({
                    'producto': detalle_pedido.product.name,
                    'cantidad': detalle_pedido.quantity,
                    'precio': detalle_pedido.price,
                    'subtotal': subtotal_detalle,
                })
                subtotal_pedido_actual += subtotal_detalle
                total_ingresos_productor += subtotal_detalle
        
        ventas_agrupadas.append({
            'pedido_id': pedido_actual.id,
            'fecha': pedido_actual.created_at,
            'comprador': pedido_actual.user.get_full_name() or pedido_actual.user.username,
            'items': items_del_productor_en_pedido,
            'total_items': len(items_del_productor_en_pedido),
            'subtotal_pedido': subtotal_pedido_actual,
        })
    
    context = {
        'productor': productor_seleccionado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'ventas_agrupadas': ventas_agrupadas,
        'total_ingresos': total_ingresos_productor,
        'total_ventas': total_ventas_productor,
    }
    return render(request, 'complaints/reporte_ventas_productor.html', context)


# =============================================================================
# FUNCIÓN: reporte_fincas_excel
# =============================================================================
# Propósito: Exporta a Excel un reporte detallado de todas las fincas (productores)
# con información de sus productos, stock y ventas.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Archivo Excel con el reporte de fincas
#
# URLs asociadas:
#   - /denuncias/reportes/fincas/excel/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def reporte_fincas_excel(request):
    """Exporta a Excel el reporte de fincas con productos, stock y stock disponible"""
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_fincas.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fincas y Productos'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    
    headers = ['Nombre', 'Cédula', 'Finca/Negocio', 'Teléfono', 'Email', 'Producto', 'Categoría', 'Precio', 'Stock Inicial', 'Vendido', 'Stock Disponible']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    productores_activos = Usuario.objects.filter(role='productor', is_approved=True)
    fila_actual = 2
    
    for productor_actual in productores_activos:
        productos_vendedor = Producto.objects.filter(vendedor=productor_actual)
        if productos_vendedor.exists():
            for producto_actual in productos_vendedor:
                from apps.cart.models import OrderItem
                cantidad_vendida = OrderItem.objects.filter(
                    product=producto_actual,
                    order__status='delivered'
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                stock_disponible = producto_actual.stock - cantidad_vendida
                
                ws.cell(row=fila_actual, column=1, value=productor_actual.get_full_name())
                ws.cell(row=fila_actual, column=2, value=productor_actual.username)
                ws.cell(row=fila_actual, column=3, value=productor_actual.business_name or '-')
                ws.cell(row=fila_actual, column=4, value=productor_actual.phone or '-')
                ws.cell(row=fila_actual, column=5, value=productor_actual.email)
                ws.cell(row=fila_actual, column=6, value=producto_actual.name)
                ws.cell(row=fila_actual, column=7, value=producto_actual.category)
                ws.cell(row=fila_actual, column=8, value=float(producto_actual.price))
                ws.cell(row=fila_actual, column=9, value=producto_actual.stock)
                ws.cell(row=fila_actual, column=10, value=cantidad_vendida)
                ws.cell(row=fila_actual, column=11, value=stock_disponible)
                fila_actual += 1
        else:
            ws.cell(row=fila_actual, column=1, value=productor_actual.get_full_name())
            ws.cell(row=fila_actual, column=2, value=productor_actual.username)
            ws.cell(row=fila_actual, column=3, value=productor_actual.business_name or '-')
            ws.cell(row=fila_actual, column=4, value=productor_actual.phone or '-')
            ws.cell(row=fila_actual, column=5, value=productor_actual.email)
            ws.cell(row=fila_actual, column=6, value='Sin productos')
            ws.cell(row=fila_actual, column=7, value='-')
            ws.cell(row=fila_actual, column=8, value='-')
            ws.cell(row=fila_actual, column=9, value='-')
            ws.cell(row=fila_actual, column=10, value='-')
            ws.cell(row=fila_actual, column=11, value='-')
            fila_actual += 1
    
    column_widths = [20, 15, 20, 15, 25, 20, 15, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(response)
    return response


# =============================================================================
# FUNCIÓN: exportar_ventas_excel
# =============================================================================
# Propósito: Exporta a Excel el reporte general de ventas por período.
#
# Parámetros:
#   - request: Objeto HttpRequest con GET (fecha_inicio, fecha_fin)
#
# Retorna:
#   - HttpResponse: Archivo Excel con el reporte de ventas
#
# URLs asociadas:
#   - /denuncias/reportes/ventas/excel/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def exportar_ventas_excel(request):
    """Exporta a Excel el reporte general de ventas por período"""
    
    # Obtener fechas del filtro
    fecha_inicio, fecha_fin = obtener_fechas(request)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_ventas_{fecha_inicio}_{fecha_fin}.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    
    headers = ['Fecha', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal', 'Vendedor', 'Comprador']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    pedidos_periodo = Pedido.objects.filter(
        created_at__date__gte=fecha_inicio,
        created_at__date__lte=fecha_fin,
        status='delivered'
    )
    
    fila_excel = 2
    total_ingresos_excel = 0
    
    for pedido_actual in pedidos_periodo:
        for detalle_pedido in pedido_actual.items.all():
            ws.cell(row=fila_excel, column=1, value=pedido_actual.created_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=fila_excel, column=2, value=detalle_pedido.product.name)
            ws.cell(row=fila_excel, column=3, value=detalle_pedido.quantity)
            ws.cell(row=fila_excel, column=4, value=float(detalle_pedido.price))
            ws.cell(row=fila_excel, column=5, value=float(detalle_pedido.price * detalle_pedido.quantity))
            ws.cell(row=fila_excel, column=6, value=detalle_pedido.product.vendedor.get_full_name())
            ws.cell(row=fila_excel, column=7, value=pedido_actual.user.get_full_name())
            total_ingresos_excel += detalle_pedido.price * detalle_pedido.quantity
            fila_excel += 1
    
    if fila_excel > 2:
        cell = ws.cell(row=fila_excel, column=4, value='TOTAL:')
        cell.font = Font(bold=True)
        cell = ws.cell(row=fila_excel, column=5, value=float(total_ingresos_excel))
        cell.font = Font(bold=True)
    
    column_widths = [20, 25, 12, 15, 15, 25, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(response)
    return response


# =============================================================================
# FUNCIÓN: exportar_denuncias_excel
# =============================================================================
# Propósito: Exporta todas las denuncias del sistema a un archivo Excel.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Archivo Excel con todas las denuncias
#
# URLs asociadas:
#   - /denuncias/exportar-excel/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def exportar_denuncias_excel(request):
    """Exporta todas las denuncias a un archivo Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=denuncias.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Denuncias'
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    
    headers = ['Ticket', 'Título', 'Estado', 'Prioridad', 'Creado por', 'Fecha', 'Producto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    denuncias_todas = Complaint.objects.all().order_by('-created_at')
    fila_actual = 2
    
    for denuncia_actual in denuncias_todas:
        nombre_producto = 'N/A'
        if hasattr(denuncia_actual, 'producto') and denuncia_actual.producto:
            nombre_producto = denuncia_actual.producto.name
        
        ws.cell(row=fila_actual, column=1, value=denuncia_actual.ticket_number)
        ws.cell(row=fila_actual, column=2, value=denuncia_actual.title)
        ws.cell(row=fila_actual, column=3, value=denuncia_actual.get_status_display())
        ws.cell(row=fila_actual, column=4, value=denuncia_actual.get_priority_display())
        ws.cell(row=fila_actual, column=5, value=denuncia_actual.complainant.get_full_name() or denuncia_actual.complainant.username)
        ws.cell(row=fila_actual, column=6, value=denuncia_actual.created_at.strftime('%d/%m/%Y %H:%M') if denuncia_actual.created_at else '')
        ws.cell(row=fila_actual, column=7, value=nombre_producto)
        fila_actual += 1
    
    column_widths = [20, 40, 18, 15, 25, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(response)
    return response


# =============================================================================
# FUNCIÓN: exportar_consumidores_excel
# =============================================================================
# Propósito: Exporta a Excel la lista de todos los consumidores registrados.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Archivo Excel con la lista de consumidores
#
# URLs asociadas:
#   - /export/consumidores/excel/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def exportar_consumidores_excel(request):
    """Exporta a Excel la lista de todos los consumidores registrados"""
    
    from apps.users.models import User
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=consumidores.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Consumidores'
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    
    headers = ['Nombre', 'Cédula', 'Email', 'Teléfono', 'Dirección', 'Fecha de Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    consumidores_activos = User.objects.filter(role='consumidor', is_active=True).order_by('first_name', 'last_name')
    
    fila_actual = 2
    for consumidor_actual in consumidores_activos:
        ws.cell(row=fila_actual, column=1, value=consumidor_actual.get_full_name())
        ws.cell(row=fila_actual, column=2, value=consumidor_actual.username)
        ws.cell(row=fila_actual, column=3, value=consumidor_actual.email or '')
        ws.cell(row=fila_actual, column=4, value=consumidor_actual.phone or '')
        ws.cell(row=fila_actual, column=5, value=consumidor_actual.address or '')
        ws.cell(row=fila_actual, column=6, value=consumidor_actual.date_joined.strftime('%d/%m/%Y') if consumidor_actual.date_joined else '')
        fila_actual += 1
    
    column_widths = [25, 15, 30, 15, 35, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(response)
    return response


# =============================================================================
# FUNCIÓN: exportar_productores_excel
# =============================================================================
# Propósito: Exporta a Excel la lista de todos los productores registrados
# con sus métricas clave.
#
# Parámetros:
#   - request: Objeto HttpRequest
#
# Retorna:
#   - HttpResponse: Archivo Excel con la lista de productores
#
# URLs asociadas:
#   - /export/productores/excel/
#
# Roles permitidos:
#   - ✅ Regulador (ONPECO)
# =============================================================================
@onpeco_required
def exportar_productores_excel(request):
    """Exporta a Excel la lista de todos los productores registrados"""
    
    from apps.users.models import User
    from apps.marketplace.models import Product
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productores.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productores'
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    
    headers = ['Nombre', 'Cédula', 'Negocio', 'Email', 'Teléfono', 'Dirección', 'Estado', 'Total Productos', 'Reputación', 'Fecha de Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    productores_todos = User.objects.filter(role='productor').order_by('first_name', 'last_name')
    
    fila_actual = 2
    for productor_actual in productores_todos:
        total_productos_vendedor = Product.objects.filter(vendedor=productor_actual).count()
        estado_productor = 'Aprobado' if productor_actual.is_approved else 'Pendiente'
        
        ws.cell(row=fila_actual, column=1, value=productor_actual.get_full_name())
        ws.cell(row=fila_actual, column=2, value=productor_actual.username)
        ws.cell(row=fila_actual, column=3, value=productor_actual.business_name or '')
        ws.cell(row=fila_actual, column=4, value=productor_actual.email or '')
        ws.cell(row=fila_actual, column=5, value=productor_actual.phone or '')
        ws.cell(row=fila_actual, column=6, value=productor_actual.address or '')
        ws.cell(row=fila_actual, column=7, value=estado_productor)
        ws.cell(row=fila_actual, column=8, value=total_productos_vendedor)
        ws.cell(row=fila_actual, column=9, value=productor_actual.get_reputacion_display())
        ws.cell(row=fila_actual, column=10, value=productor_actual.date_joined.strftime('%d/%m/%Y') if productor_actual.date_joined else '')
        fila_actual += 1
    
    column_widths = [25, 15, 25, 30, 15, 35, 15, 15, 30, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save(response)
    return response